# -*- coding:utf8 -*-
import os
from openpyxl import load_workbook
import sys
import logging
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite_tz.settings")

django.setup()  # 必须得先设置django启动才能使用model中的删除


def node():
    from trans_book.models import Node

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理'.decode('utf8')

    wb = load_workbook(filename=os.path.join(BASE_DIR, u'新环路表样本 .xlsx'), read_only=True, data_only=True)

    ws_list = wb.get_sheet_names()  # 返回所有表名
    for ws_name in ws_list[1:]:
        print u"导%s表" % ws_name
        NodeList = []
        rows = wb[ws_name].rows  # 是一个生成器，每次调用ws.rows都会生成一个新的生成器（不知什么原因），所以需要赋值给一个变量固定下来
        rows.send(None)  # 启动生成器，表的第一行是不需要的
        for index, row in enumerate(rows):  # 除第一行外的所有行
            node_id = row[0].value
            node_name = row[12].value
            node_short_name = row[1].value
            node_type = row[3].value
            node_area = row[4].value
            node_ring = row[5].value
            chain_node = row[6].value
            NodeList.append(Node(node_id=node_id,
                                 name=node_name,
                                 short_name=node_short_name,
                                 node_type=node_type,
                                 area=node_area,
                                 ring=node_ring,
                                 chain_node=None))

        Node.objects.bulk_create(NodeList)


def chain_node():
    from trans_book.models import Node

    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('file')
    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理'.decode('utf8')

    wb = load_workbook(filename=os.path.join(BASE_DIR, u'新环路表样本 .xlsx'), read_only=True, data_only=True)

    ws_list = wb.get_sheet_names()
    for ws_name in ws_list[1:]:
        print u'导入%s表的上联节点' % ws_name
        rows = wb[ws_name].rows
        rows.send(None)
        for index, row in enumerate(rows):
            chain_node = row[6].value
            node_name = row[1].value
            ring = row[5].value
            if chain_node != None:
                print u"%s的上联节点：%s" % (row[1].value, chain_node)
                try:
                    node = Node.objects.get(short_name=node_name, ring=ring)  # 如何选取节点
                except Exception, e:
                    message = u"%s表%s行--获取主节点\"%s\"出错\n%s" % (ws_name, index + 2, node_name, str(e))
                    message = message.encode('gbk')
                    logger.error(message)
                try:
                    node.chain_node = Node.objects.get(short_name=chain_node, ring=ring)
                    node.save()  # 得save才会更新
                except Exception, e:

                    message = u"%s表%s行--获取\"%s\"的上联节点\"%s\"出错\n%s" % (ws_name, index + 2, node_name, chain_node, str(e))

                    message = message.encode('gbk')
                    # f.write(str(e))
                    # f.write('\n')
                    logger.error(message)


def import_hw_3g_sf():
    from trans_book.models import hw_3g_sf, Node

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理'.decode('utf8')

    wb = load_workbook(filename=os.path.join(BASE_DIR, u'副本3G基站电路台帐(最新).xlsm'), read_only=True, data_only=True)

    ws = wb[u'华为3G室分']
    # ws_list = wb.get_sheet_names()  #返回所有表名
    # for ws_name in ws_list[1:]:
    # print u"导%s表" % ws_name
    List = []
    f = open('import_hw_3g_sf.txt', 'w')
    rows = ws.rows  # 是一个生成器，每次调用ws.rows都会生成一个新的生成器（不知什么原因），所以需要赋值给一个变量固定下来
    rows.send(None)  # 启动生成器，表的第一行是不需要的
    for index, row in enumerate(rows):  # 除第一行外的所有行
        jz_name = row[0].value
        jz_route = row[6].value
        source_node = row[20].value
        try:
            node = Node.objects.get(name=source_node)
        except Exception, e:
            node = None
            f.write('%s\n' % str(e))
            # f.write(jz_name.encode('gbk'))
            f.write('\n')
        List.append(hw_3g_sf(jz_name=jz_name,
                             jz_route=jz_route,
                             source_node=node))

    hw_3g_sf.objects.bulk_create(List)
    f.close()


def import_hw_3g_jz():
    from trans_book.models import hw_3g_jz, Node

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理'.decode('utf8')

    wb = load_workbook(filename=os.path.join(BASE_DIR, u'副本3G基站电路台帐(最新).xlsm'), read_only=True, data_only=True)

    ws = wb[u'华为3G宏蜂窝']
    # ws_list = wb.get_sheet_names()  #返回所有表名
    # for ws_name in ws_list[1:]:
    # print u"导%s表" % ws_name
    f = open('import_hw_3g_jz.txt', 'w')
    List = []
    rows = ws.rows  # 是一个生成器，每次调用ws.rows都会生成一个新的生成器（不知什么原因），所以需要赋值给一个变量固定下来
    rows.send(None)  # 启动生成器，表的第一行是不需要的
    for index, row in enumerate(rows):  # 除第一行外的所有行
        jz_name = row[0].value
        jz_route = row[6].value
        source_node = row[19].value
        try:
            node = Node.objects.get(name=source_node)
        except Exception, e:
            node = None
            f.write('%s\n' % str(e))
            if jz_name == None:
                jz_name = 'None'
            f.write(jz_name.encode('gbk'))
            f.write('\n')
        List.append(hw_3g_jz(jz_name=jz_name,
                             jz_route=jz_route,
                             source_node=node))

    hw_3g_jz.objects.bulk_create(List)


def import_hw_3g_fe():
    from trans_book.models import hw_3g_fe, Node

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理'.decode('utf8')

    wb = load_workbook(filename=os.path.join(BASE_DIR, u'副本3G基站电路台帐(最新).xlsm'), read_only=True, data_only=True)

    ws = wb[u'华为FE电路']
    # ws_list = wb.get_sheet_names()  #返回所有表名
    # for ws_name in ws_list[1:]:
    # print u"导%s表" % ws_name
    List = []
    f = open('import_hw_3g_fe.txt', 'w')
    rows = ws.rows  # 是一个生成器，每次调用ws.rows都会生成一个新的生成器（不知什么原因），所以需要赋值给一个变量固定下来
    rows.send(None)  # 启动生成器，表的第一行是不需要的
    for index, row in enumerate(rows):  # 除第一行外的所有行
        jz_name = row[1].value
        jz_route = row[6].value
        source_node_short = row[17].value
        source_node = row[20].value
        if source_node_short != None:
            try:
                node = Node.objects.get(name=source_node)
            except Exception, e:
                node = None
                f.write('%s\n' % str(e))

                # f.write(jz_name.encode('gbk'))
                f.write('\n\n')
            List.append(hw_3g_fe(jz_name=jz_name,
                                 jz_route=jz_route,
                                 source_node=node))

    hw_3g_fe.objects.bulk_create(List)


"""
def import_hw_3g_sf_source():
    from trans_book.models import hw_3g_sf

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理'.decode('utf8')

    wb = load_workbook(filename = os.path.join(BASE_DIR, u'副本3G基站电路台帐(最新).xlsm'), read_only=True, data_only=True)

    ws = wb[u'华为3G室分']

    rows = ws.rows
    rows.send(None)

    for index, row in enumerate(rows):
        source_node = row[18].value
        if souerce_node!= None:
            try:
                node = hw_3g_sf.objects.get(short_name=node_name)  #如何选取节点
                node.chain_node = Node.objects.get(short_name=chain_node)
                node.save()
            except Exception, e:
                f = open('record.txt','a')
                str1 = u"%s的上联节点：%s" % (row[1].value, chain_node)
                f.write(str1.encode('gbk'))
                f.write('\n')
                f.write(str(e))
                f.write('\n')
"""


def test():
    from trans_book.models import hw_ipran_link, hw_ipran_ring

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')

    import xlrd, re

    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'纤缆连接关系表_02-19-2016_14-50-01.xls'))
    ws = wb.sheets()[0]  # 第一张表
    LinkList = []

    pattern = re.compile(r"R\d{3,4}")  # 用于提取环路名称

    # import pdb;pdb.set_trace()
    for rownum in range(8, ws.nrows):
        row = ws.row_values(rownum)
        source = row[5]  # 源
        dest = row[7]  # 宿

        ringName1 = pattern.search(source)
        ringName2 = pattern.search(dest)

        if ringName1:
            ring = ringName1.group()
        elif ringName2:
            ring = ringName2.group()
        else:
            ring = None

        if ring:
            link_obj = hw_ipran_link(
                    link=(source, dest),
                    ring=hw_ipran_ring.objects.get(name=ring)
            )
        else:
            link_obj = hw_ipran_link(
                    link=(source, dest)
            )

        LinkList.append(link_obj)
    hw_ipran_link.objects.bulk_create(LinkList)


def generate_search_table_by_node_table():
    from trans_book.models import Node, NodeSearch

    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('Chain')

    import time
    t1 = time.time()
    List = []
    for row in Node.objects.all():
        if row.chain_node == None:
            parent_id = None
        else:
            parent_id = row.chain_node.id
        List.append(NodeSearch(
                id=row.id,
                parent_id=parent_id,
                node_path=str(row.id)))
    NodeSearch.objects.bulk_create(List)
    print time.time() - t1

    for row in NodeSearch.objects.all():  # 操作父节点不为空的点，并生成节点路径：A-B-C(A，B为父节点)，存入数据库
        a = row.parent_id
        if a != None:
            message = u"节点pk:%s，父节点pk:%s" % (row.id, row.parent_id)
            logger.debug(message.encode('gbk'))

            b = row.node_path

            i = 0
            while a != None:
                if i > 30:
                    message = u"循环状态,强制跳出。"
                    logger.debug(message.encode('gbk'))
                    break
                i += 1
                b = str(a) + '-' + b
                a = NodeSearch.objects.get(id=a).parent_id
                message = u"父节点pk:%s,节点路径:%s" % (a, b)
                logger.debug(message.encode('gbk'))

            row.node_path = b
            row.save()
    print time.time() - t1

    for row in NodeSearch.objects.all():  # 根据某个节点搜索上面生成的节点路径，并得出下挂节点数
        a = row.node_path
        b = NodeSearch.objects.filter(node_path__contains=a).count()

        if b > 1:
            message = u"节点路径：%s，下挂节点数：%d" % (a, b)
            logger.debug(message.encode('gbk'))

            row.leaves_nums = b
            c = NodeSearch.objects.filter(node_path__contains=a)
            leaves = [c_row.id for c_row in c]
            row.leaves = leaves
            row.save()

    print time.time() - t1  # 总计38.5S，因为使用count并不需要花很多时间，只是返回条数，没有提取值。

    # import pdb
    # pdb.set_trace()
    """
    for row in NodeSearch.objects.all():
        a = row.node_path



        if row.leaves_nums>1:
            b = NodeSearch.objects.filter(node_path__contains=a)
            print a,b
            leaves = [b_row.id for b_row in b]
            #row.leaves_nums = len(b)
            row.leaves = leaves
            row.save()



    print time.time()-t1    #总计40S，


    for row in NodeSearch.objects.all():
        a = row.node_path
        b = NodeSearch.objects.filter(node_path__contains=a).count()
        print a,b
        if b>1:

            row.leaves_nums = b
            row.save()

    print time.time()-t1    
    """


def get_down_nodes(self, node):  # 先获取下挂节点的函数
    # import pdb
    # pdb.set_trace()
    node_down = node.node_set.all()  # 结果是一个listzNone
    node_down_list = []
    business_down_list = []
    node_down_list.append({node.pk: node.name})

    business_down_list.extend(self.get_node_business(node))


def remove_all(table):
    from trans_book.models import *
    # import pdb;pdb.set_trace()
    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('File')
    # globals()[table].objects.all().delete()
    locals()[table].objects.all().delete()

    print u'删除%s表成功！' % table


def import_hw_ipran_ring():  # 从环路表中导入环路及对应的带环点
    from trans_book.models import hw_ipran_ring

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')

    import xlrd, re

    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'分组网环路结构.xlsx'))
    ws = wb.sheets()[0]  # 第四张表

    ringList = []
    for rownum in range(1, ws.nrows):
        row = ws.row_values(rownum)
        name = row[0]   #第一列名称
        root = row[1]  #节点名
        if root!=None:
            rootList = root.split('|')

            ringObj = hw_ipran_ring(
                    name = name,
                    root = rootList
            )
            ringList.append(ringObj)

    hw_ipran_ring.objects.bulk_create(ringList)



    """
    RingList = []
    RingDict = {}

    pattern1 = re.compile(r'^[RH]\d+')    #用于匹配环路名，例如R001、H001
    pattern2 = re.compile(r'(.*?)( |XGE|GE).*')   #用于匹配ASG名称，不提取端口

    #import pdb;pdb.set_trace()
    for rownum in range(1,ws.nrows):
        row  = ws.row_values(rownum)
        #name = row[2]   #环路名称，第3列
        name = pattern1.match(row[2])    #环路名称，第3列
        if name:
            name = name.group()

        asg  = pattern2.match(row[1])    #ASG名称
        if asg:
            asg = asg.group(1)

        if name==None:
            pass
        elif name in RingList:
            #RingDict[name].append(repr(asg))
            RingDict[name].append(asg)  #数据库中存的数据是unicode编码，而不是中文
        else:
            RingList.append(name)
            RingDict[name] = [asg]
    pass

    tmp = []
    for i in RingList:
        tmp.append(hw_ipran_ring(
            name        = i,
            root        = RingDict[i]
        ))
    hw_ipran_ring.objects.bulk_create(tmp)
    """


def import_hw_ipran_ring_abandon():  # 从环路表中导入环路及对应的带环点，作废，改为自己的表来导入。16.2.25
    from trans_book.models import hw_ipran_ring

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')

    import xlrd, re

    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'ipran环路命名表.xlsm'))
    ws = wb.sheets()[3]  # 第四张表
    RingList = []
    RingDict = {}

    pattern1 = re.compile(r'^[RH]\d+')  # 用于匹配环路名，例如R001、H001
    pattern2 = re.compile(r'(.*?)( |XGE|GE).*')  # 用于匹配ASG名称，不提取端口

    # import pdb;pdb.set_trace()
    for rownum in range(1, ws.nrows):
        row = ws.row_values(rownum)
        # name = row[2]   #环路名称，第3列
        name = pattern1.match(row[2])  # 环路名称，第3列
        if name:
            name = name.group()

        asg = pattern2.match(row[1])  # ASG名称
        if asg:
            asg = asg.group(1)

        if name == None:
            pass
        elif name in RingList:
            # RingDict[name].append(repr(asg))
            RingDict[name].append(asg)  # 数据库中存的数据是unicode编码，而不是中文
        else:
            RingList.append(name)
            RingDict[name] = [asg]
    pass

    tmp = []
    for i in RingList:
        tmp.append(hw_ipran_ring(
                name=i,
                root=RingDict[i]
        ))
    hw_ipran_ring.objects.bulk_create(tmp)


def import_hw_ipran_node():  # 导入华为分组网网元节点
    from trans_book.models import hw_ipran_node, hw_ipran_ring
    import converter

    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('file')
    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')

    import xlrd

    #pattern1 = re.compile(r"[HRJ]\d{3,4}")      #用于提取网元名中的环路
    #pattern2 = re.compile(r"[\一-\龥]+")
    #pattern2 = re.compile(r'^[\-a-zA-Z0-9 ]+')    #用于清除网元名中的前缀的英文字母及连接符,空格
    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'网元信息报表_03-04-2016_17-00-18.xls'))
    ws = wb.sheets()[0]
    NodeList = []
    for rownum in range(8, ws.nrows):
        row = ws.row_values(rownum)
        name = row[0]  # 网元全称
        #ShortName = pattern2.sub('',row[0],1)   #只清除一次，网元名中的前缀的英文字母及连接符
        ShortName = converter.hw_ipran.name_convert(name)
        node_type = row[1]
        ip = row[2]
        #RingName = pattern1.search(name)    #提取环路信息
        RingName = converter.hw_ipran.access_ring_filter(name)

        if RingName is None:
            message = u'节点:%s查无环路编号' % name
            message = message.encode('gbk')
            logger.info(message)
            RingObj = None
        else:
            #RingName = RingName.group()
            try:
                RingObj = hw_ipran_ring.objects.get(name=RingName)
            except:
                message = u'环路表查无%s' % RingName
                message = message.encode('gbk')
                logger.error(message)
                RingObj = None
        """
        node = hw_ipran_node(
                name=name,
                ShortName=ShortName,
                node_type=node_type,
                ip=ip
                ring=RingObj
        )

        
        node.save()

        
        if RingName is None:
            message = u'节点:%s查无环路编号' % name
            message = message.encode('gbk')
            logger.info(message)
        else:
            RingName = RingName.group()
            try:
                ring = hw_ipran_ring.objects.get(name=RingName)
                node.ring.add(ring)
            except:
                message = u'环路表查无%s' % RingName
                message = message.encode('gbk')
                logger.error(message)

        """
        NodeList.append(hw_ipran_node(
            name        = name,
            node_type   = node_type,
            ip          = ip,
            ShortName  = ShortName,
            ring        = RingObj
        ))
    hw_ipran_node.objects.bulk_create(NodeList)

    """
    导入节点名称后，接下来导入所属环路，无法一次导入，因为得调用主键来创建中间表
    """


def import_hw_ipran_link():  # 导入华为分组网的网元对接关系
    from trans_book.models import hw_ipran_link, hw_ipran_ring
    import logging

    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('file')

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')

    import xlrd
    import converter

    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'纤缆连接关系表_03-04-2016_16-58-28.xls'))
    ws = wb.sheets()[0]  # 第一张表
    LinkList = []

    #pattern = re.compile(r"R\d{3,4}")  # 用于提取环路名称

    # import pdb;pdb.set_trace()
    for rownum in range(8, ws.nrows):
        row = ws.row_values(rownum)
        source = row[5]  # 源
        dest = row[7]  # 宿
        sport = row[6]  #源端口
        dport = row[8]  #宿端口

        ringName1 = converter.hw_ipran.access_ring_filter(source)
        ringName2 = converter.hw_ipran.access_ring_filter(dest)

        if ringName1:
            ring = ringName1
        elif ringName2:
            ring = ringName2
        else:
            ring = None

        if ring:
            try:
                RingObj = hw_ipran_ring.objects.get(name=ring)
            except:
                message = u'查无该环路:%s' % ring
                message = message.encode('gbk')
                logger.error(message)
                RingObj = None
            link_obj = hw_ipran_link(
                    #link=(source, dest),
                    source = source,
                    dest   = dest,
                    ring   = RingObj
            )
        else:
            link_obj = hw_ipran_link(
                    source = source,
                    dest   = dest,
            )

        LinkList.append(link_obj)
    hw_ipran_link.objects.bulk_create(LinkList)

def import_hw_ipran_business_2G():
    import logging
    from trans_book.models import hw_ipran_node, hw_ipran_ring, hw_ipran_business_2G
    import xlrd
    import converter

    #pattern = re.compile(r'^[\-a-zA-Z0-9 ]+')    #用于清除网元名中的前缀的英文字母及连接符，空格

    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('file')

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')
    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'IPRAN业务台账.xlsm'))
    ws = wb.sheets()[0]  # 第一张表
    
    BusinessList = []
    for index in range(1, ws.nrows):
        row = ws.row_values(index)
        rownum = index+1

        StationName = row[0]

        TransStationName = row[1]     #传输自编号
        if TransStationName=="":        #传输站名为空的话就不导入这一行
            message = u'%d行传输自编号为空' % rownum
            message = message.encode('gbk')
            logger.info(message)
            continue    #传输自编号为空则跳过这一行，不录入数据库

        #TransName = row[7]
        #TransName = pattern.sub("",row[7],1)
        TransName = converter.hw_ipran.name_convert(row[7])
        try:    
            Node = hw_ipran_node.objects.get(ShortName=TransName)
        except Exception:
            message = u'%d行查无匹配节点:%s' % (rownum,TransName)
            message = message.encode('gbk')
            logger.error(message)
            Node = None

        try:
            Ring = hw_ipran_ring.objects.get(name=row[6])
        except Exception:
            message = u'%d行查无匹配环路:%s'  % (rownum,row[6])
            message = message.encode('gbk')
            logger.error(message)
            Ring = None

        BusinessObj = hw_ipran_business_2G(
                    StationName = StationName,
                    TransStationName = TransStationName,
                    Node = Node,
                    Ring = Ring
            )
        BusinessList.append(BusinessObj)

    hw_ipran_business_2G.objects.bulk_create(BusinessList)

def import_hw_ipran_business_3G_E1():
    import logging
    from trans_book.models import hw_ipran_node, hw_ipran_ring, hw_ipran_business_3G_E1
    import xlrd
    import converter

    #pattern = re.compile(r'^[\-a-zA-Z0-9 ]+')    #用于清除网元名中的前缀的英文字母及连接符

    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('file')

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')
    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'IPRAN业务台账.xlsm'))
    ws = wb.sheets()[1]  # 第er张表
    
    BusinessList = []
    for index in range(1, ws.nrows):
        row = ws.row_values(index)
        rownum = index+1

        StationName = row[0]

        TransStationName = row[1]     #传输自编号
        if TransStationName=="":        #传输站名为空的话就不导入这一行
            message = u'%d行传输自编号为空' % rownum
            message = message.encode('gbk')
            logger.info(message)
            continue    #传输自编号为空则跳过这一行，不录入数据库

        #TransName = row[7]
        #TransName = pattern.sub("",row[7],1)
        #TransName = name_converter.hw_ipran.convert(row[7])
        TransName = converter.hw_ipran.name_convert(row[7])
        try:
            Node = hw_ipran_node.objects.get(ShortName=TransName)
        except Exception:
            message = u'%d行查无匹配节点:%s' % (rownum,TransName)
            message = message.encode('gbk')
            logger.error(message)
            Node = None

        try:
            Ring = hw_ipran_ring.objects.get(name=row[6])
        except Exception:
            message = u'%d行查无匹配环路:%s'  % (rownum,row[6])
            message = message.encode('gbk')
            logger.error(message)
            Ring = None

        BusinessObj = hw_ipran_business_3G_E1(
                    StationName = StationName,
                    TransStationName = TransStationName,
                    Node = Node,
                    Ring = Ring
            )
        BusinessList.append(BusinessObj)

    hw_ipran_business_3G_E1.objects.bulk_create(BusinessList)

def import_hw_ipran_business_3G_FE():
    import logging
    from trans_book.models import hw_ipran_node, hw_ipran_ring, hw_ipran_business_3G_FE
    import xlrd
    import converter

    #pattern = re.compile(r'^[\-a-zA-Z0-9 ]+')    #用于清除网元名中的前缀的英文字母及连接符

    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('file')

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')
    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'IPRAN业务台账.xlsm'))
    ws = wb.sheets()[2]  # 第三张表
    
    BusinessList = []
    for index in range(1, ws.nrows):
        row = ws.row_values(index)
        rownum = index+1

        StationName = row[0]

        TransStationName = row[1]     #传输自编号
        if TransStationName=="":        #传输站名为空的话就不导入这一行
            message = u'%d行传输自编号为空' % rownum
            message = message.encode('gbk')
            logger.info(message)
            continue    #传输自编号为空则跳过这一行，不录入数据库

        #TransName = row[7]
        #TransName = pattern.sub("",row[7],1)
        TransName = converter.hw_ipran.name_convert(row[7])
        try:
            Node = hw_ipran_node.objects.get(ShortName=TransName)
        except Exception:
            message = u'%d行查无匹配节点:%s' % (rownum,TransName)
            message = message.encode('gbk')
            logger.error(message)
            Node = None

        try:
            Ring = hw_ipran_ring.objects.get(name=row[6])
        except Exception:
            message = u'%d行查无匹配环路:%s'  % (rownum,row[6])
            message = message.encode('gbk')
            logger.error(message)
            Ring = None

        BusinessObj = hw_ipran_business_3G_FE(
                    StationName = StationName,
                    TransStationName = TransStationName,
                    Node = Node,
                    Ring = Ring
            )
        BusinessList.append(BusinessObj)

    hw_ipran_business_3G_FE.objects.bulk_create(BusinessList)

def import_hw_ipran_business_LTE():
    import logging
    from trans_book.models import hw_ipran_node, hw_ipran_ring, hw_ipran_business_LTE
    import xlrd
    import converter

    #pattern = re.compile(r'^[\-a-zA-Z0-9 ]+')    #用于清除网元名中的前缀的英文字母及连接符

    logger = logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('file')
    #logger.addHandler(logger.getHandler('fileHandler','trys.txt'))

    BASE_DIR = r'C:\Users\Administrator\Desktop\环路表整理\华为分组网'.decode('utf8')
    wb = xlrd.open_workbook(os.path.join(BASE_DIR, u'IPRAN业务台账.xlsm'))
    ws = wb.sheets()[3]  # 第四张表
    
    BusinessList = []
    for index in range(1, ws.nrows):
        row = ws.row_values(index)
        rownum = index+1

        StationName = row[0]

        TransStationName = row[1]     #传输自编号
        if TransStationName=="":        #传输站名为空的话就不导入这一行
            message = u'%d行传输自编号为空' % rownum
            message = message.encode('gbk')
            logger.info(message)
            continue    #传输自编号为空则跳过这一行，不录入数据库

        #TransName = row[7]
        #TransName = pattern.sub("",row[7],1)
        TransName = converter.hw_ipran.name_convert(row[7])
        try:
            Node = hw_ipran_node.objects.get(ShortName=TransName)
        except Exception:
            message = u'%d行查无匹配节点:%s' % (rownum,TransName)
            message = message.encode('gbk')
            logger.error(message)
            Node = None

        try:
            Ring = hw_ipran_ring.objects.get(name=row[6])
        except Exception:
            message = u'%d行查无匹配环路:%s'  % (rownum,row[6])
            message = message.encode('gbk')
            logger.error(message)
            Ring = None

        BusinessObj = hw_ipran_business_LTE(
                    StationName = StationName,
                    TransStationName = TransStationName,
                    Node = Node,
                    Ring = Ring
            )
        BusinessList.append(BusinessObj)

    message = u'总计录入%d行数据' % rownum
    message = message.encode('gbk')
    logger.info(message)
    hw_ipran_business_LTE.objects.bulk_create(BusinessList)


def main(argv):

    param1 = argv[1]
    if param1 == 'node':
        node()
    elif param1 == 'chain':
        chain_node()
    elif param1 == 'jz':
        import_hw_3g_jz()
    elif param1 == 'sf':
        import_hw_3g_sf()
    elif param1 == 'fe':
        import_hw_3g_fe()
    elif param1 == 'test':
        test()
    elif param1 == 'a1':
        import_hw_ipran_ring()
    elif param1 == 'b1':
        import_hw_ipran_node()
    elif param1 == 'c1':
        import_hw_ipran_link()
    elif param1 == 'd1':
        import_hw_ipran_business_2G()
    elif param1 == 'd2':
        import_hw_ipran_business_3G_E1()
    elif param1 == 'd3':
        import_hw_ipran_business_3G_FE()
    elif param1 == 'd4':
        import_hw_ipran_business_LTE()
    elif param1 == 'remove':
        param2 = argv[2]
        remove_all(param2)


if __name__ == "__main__":
    main(sys.argv)  # sys.argv是一个列表
    # chain_node()    #导入节点的上联节点
    # import_hw_3g_sf()   #导入华为3G室分表
    # import_hw_3g_jz()
    # import_hw_3g_fe()
    # import_hw_3g_sf_source()

    print 'Done!'
